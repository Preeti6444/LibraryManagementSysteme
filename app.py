<<<<<<< HEAD
from flask import Flask, request, render_template, redirect, url_for, jsonify
import pandas as pd

app = Flask(__name__)

@app.route('/')
def display():
    df = pd.read_excel('Library.xlsx')
    blogs = df.to_dict('records')
    return render_template('display.html', blogs=blogs)

@app.route('/search')
def search_book():
    title = request.args.get('title', '').strip().lower()
    df = pd.read_excel('Library.xlsx')

    # Normalize title column
    df['title'] = df['title'].astype(str).str.strip().str.lower()

    result_text = "Not Available"
    matching = df[df['title'] == title]

    if not matching.empty:
        status = matching.iloc[0]['status'].strip().lower()
        if status == 'available':
            result_text = "Available"
        elif status == 'issued':
            result_text = " Issued"
        else:
            result_text = f" Status: {status}"

    blogs = df.to_dict('records')
    return render_template('display.html', blogs=blogs, result=result_text)

@app.route('/add')
def add_blog():
    return render_template('add.html')

@app.route('/show')
def show():
    df = pd.read_excel('Library.xlsx')
    blogs = df.to_dict('records')
    return jsonify(blogs)

@app.route('/create', methods=['POST'])
def create_blog():
    title = request.form.get('title')
    author = request.form.get('author')
    status = request.form.get('status')
    df = pd.read_excel('Library.xlsx')
    new_id = df['id'].max() + 1 if not df.empty else 1
    df.loc[len(df)] = [new_id, title, author, status]
    df.to_excel('Library.xlsx', index=False)
    return redirect(url_for('display'))

@app.route('/delete/<string:title>')
def delete(title):
    df = pd.read_excel('Library.xlsxx')
    df = df[df['title'].str.upper().str.strip() != title.upper().strip()]
    df.to_excel('Library.xlsx', index=False)
    return redirect(url_for('display'))

@app.route('/edit/<int:blog_id>')
def edit_blog(blog_id):
    df = pd.read_excel('Library.xlsx')
    blog_row = df[df['id'] == blog_id]
    if blog_row.empty:
        return "Book not found", 404
    blog = blog_row.to_dict('records')[0]
    return render_template('edit.html', blog=blog)

@app.route('/api/update/<int:blog_id>', methods=['PUT'])
def api_update_blog(blog_id):
    df = pd.read_excel('Library.xlsx')
    if blog_id in df['id'].values:
        data = request.get_json()
        df.loc[df['id'] == blog_id, 'title'] = data.get('title')
        df.loc[df['id'] == blog_id, 'author'] = data.get('author')
        df.loc[df['id'] == blog_id, 'status'] = data.get('status')
        df.to_excel('Library.xlsx', index=False)
        return jsonify({"message": "Book updated"}), 200
    else:
        return jsonify({"error": "Book not found"}), 404

@app.route('/api/delete/<int:blog_id>', methods=['DELETE'])
def api_delete_blog(blog_id):
    df = pd.read_excel('Library.xlsx')
    if blog_id in df['id'].values:
        df = df[df['id'] != blog_id]
        df.to_excel('Library.xlsx', index=False)
        return jsonify({"message": "Book deleted"}), 200
    else:
        return jsonify({"error": "Book not found"}), 404

@app.route('/insert', methods=['POST'])
def insert():
    data = request.get_json()
    title = data['title']
    author = data['author']
    status = data['status']
    df = pd.read_excel('Library.xlsx')
    new_id = df['id'].max() + 1 if not df.empty else 1
    df.loc[len(df)] = [new_id, title, author, status]
    df.to_excel('Library.xlsx', index=False)
    return jsonify({'Message': 'Success...Data Inserted'})

if __name__ == '__main__':
    app.run(debug=True)
=======
from flask import Flask, render_template, request, redirect, url_for
import openpyxl

app = Flask(__name__)
EXCEL_FILE = 'Library.xlsx'

def load_books():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    books = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        books.append({
            'Book ID': row[0],
            'Title': row[1],
            'Author': row[2],
            'Status': row[3]
        })
    return books

def save_books(books):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Book ID', 'Title', 'Author', 'Status'])
    for book in books:
        ws.append([book['Book ID'], book['Title'], book['Author'], book['Status']])
    wb.save(EXCEL_FILE)

@app.route('/')
def home():
    books = load_books()
    available = sum(1 for b in books if b['Status'] == 'Available')
    issued = sum(1 for b in books if b['Status'] == 'Issued')
    return render_template('home.html', available=available, issued=issued)

@app.route('/books')
def book_list():
    books = load_books()
    return render_template('books.html', books=books)

@app.route('/issue', methods=['GET', 'POST'])
def issue_book():
    if request.method == 'POST':
        book_id = request.form['book_id']
        books = load_books()
        for book in books:
            if book['Book ID'] == book_id and book['Status'] == 'Available':
                book['Status'] = 'Issued'
                save_books(books)
                return redirect(url_for('book_list'))
        return "Book not available or already issued."
    return render_template('issue.html')

@app.route('/return', methods=['GET', 'POST'])
def return_book():
    if request.method == 'POST':
        book_id = request.form['book_id']
        books = load_books()
        for book in books:
            if book['Book ID'] == book_id and book['Status'] == 'Issued':
                book['Status'] = 'Available'
                save_books(books)
                return redirect(url_for('book_list'))
        return "Book not found or already returned."
    return render_template('return.html')

if __name__ == '__main__':
    app.run(debug=True)
>>>>>>> 43d847e3bbd53dbb3646e6181d2855425f7651e4
